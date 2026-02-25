#!/usr/bin/env python3
"""
Arena Protocol — Institutional-Grade Financial Model
Morgan Stanley / Goldman Sachs quality formatting
6 tabs, all formulas reference Assumptions tab
"""

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from copy import copy

# ═══════════════════════════════════════════════════
# STYLE DEFINITIONS
# ═══════════════════════════════════════════════════

NAVY = "1B2A4A"
LIGHT_NAVY = "2C3E6B"
LIGHT_GRAY = "F2F2F2"
MED_GRAY = "D9D9D9"
WHITE = "FFFFFF"
INPUT_BLUE = "0066CC"
FORMULA_BLACK = "000000"
ACCENT_GREEN = "1A7A3A"
ACCENT_RED = "C0392B"
BORDER_GRAY = "B0B0B0"

FONT_CALIBRI = "Calibri"

# Fonts
font_header = Font(name=FONT_CALIBRI, bold=True, color=WHITE, size=11)
font_section = Font(name=FONT_CALIBRI, bold=True, color=NAVY, size=11)
font_subsection = Font(name=FONT_CALIBRI, bold=True, color=LIGHT_NAVY, size=10)
font_input = Font(name=FONT_CALIBRI, color=INPUT_BLUE, size=10)
font_input_bold = Font(name=FONT_CALIBRI, bold=True, color=INPUT_BLUE, size=10)
font_formula = Font(name=FONT_CALIBRI, color=FORMULA_BLACK, size=10)
font_formula_bold = Font(name=FONT_CALIBRI, bold=True, color=FORMULA_BLACK, size=10)
font_total = Font(name=FONT_CALIBRI, bold=True, color=NAVY, size=11)
font_title = Font(name=FONT_CALIBRI, bold=True, color=WHITE, size=14)
font_subtitle = Font(name=FONT_CALIBRI, italic=True, color="666666", size=9)
font_label = Font(name=FONT_CALIBRI, color="333333", size=10)
font_pct_green = Font(name=FONT_CALIBRI, color=ACCENT_GREEN, size=10)
font_pct_red = Font(name=FONT_CALIBRI, color=ACCENT_RED, size=10)
font_white_bold = Font(name=FONT_CALIBRI, bold=True, color=WHITE, size=10)

# Fills
fill_navy = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
fill_light_navy = PatternFill(start_color=LIGHT_NAVY, end_color=LIGHT_NAVY, fill_type="solid")
fill_light_gray = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")
fill_med_gray = PatternFill(start_color=MED_GRAY, end_color=MED_GRAY, fill_type="solid")
fill_white = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")
fill_input = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")

# Borders
thin_side = Side(style="thin", color=BORDER_GRAY)
thick_side = Side(style="medium", color=NAVY)
double_side = Side(style="double", color=NAVY)
border_thin = Border(bottom=thin_side)
border_bottom_thick = Border(bottom=thick_side)
border_top_double = Border(top=double_side, bottom=thick_side)
border_all_thin = Border(top=thin_side, bottom=thin_side, left=thin_side, right=thin_side)

# Alignments
align_left = Alignment(horizontal="left", vertical="center")
align_right = Alignment(horizontal="right", vertical="center")
align_center = Alignment(horizontal="center", vertical="center")
align_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)

# Number formats
NUM_INT = '#,##0'
NUM_DOLLAR = '$#,##0'
NUM_DOLLAR_2 = '$#,##0.00'
NUM_PCT = '0.0%'
NUM_PCT_1 = '0.0%'
NUM_BPS = '#,##0" BPS"'


def apply_style(cell, font=None, fill=None, alignment=None, border=None, number_format=None):
    if font: cell.font = font
    if fill: cell.fill = fill
    if alignment: cell.alignment = alignment
    if border: cell.border = border
    if number_format: cell.number_format = number_format


def style_header_row(ws, row, max_col):
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        apply_style(cell, font=font_header, fill=fill_navy, alignment=align_center)


def style_section_header(ws, row, label, max_col):
    ws.cell(row=row, column=1, value=label)
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        apply_style(cell, font=font_section, fill=fill_light_gray, border=border_bottom_thick)


def style_total_row(ws, row, max_col):
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        apply_style(cell, font=font_total, border=border_top_double)


def col(n):
    """1-indexed column to letter."""
    return get_column_letter(n)


# ═══════════════════════════════════════════════════
# WORKBOOK CREATION
# ═══════════════════════════════════════════════════

wb = openpyxl.Workbook()

# ═══════════════════════════════════════════════════
# TAB 1: ASSUMPTIONS
# ═══════════════════════════════════════════════════

ws_a = wb.active
ws_a.title = "Assumptions"
ws_a.sheet_properties.tabColor = NAVY

# Disable gridlines
ws_a.sheet_view.showGridLines = False

# Column widths
ws_a.column_dimensions['A'].width = 4
ws_a.column_dimensions['B'].width = 42
ws_a.column_dimensions['C'].width = 18
ws_a.column_dimensions['D'].width = 18
ws_a.column_dimensions['E'].width = 18
ws_a.column_dimensions['F'].width = 18

# Title row
ws_a.merge_cells('A1:F1')
c = ws_a.cell(row=1, column=1, value="THE ARENA PROTOCOL — MODEL ASSUMPTIONS")
apply_style(c, font=font_title, fill=fill_navy, alignment=align_center)
for cc in range(2, 7):
    apply_style(ws_a.cell(row=1, column=cc), fill=fill_navy)

ws_a.merge_cells('A2:F2')
c = ws_a.cell(row=2, column=1, value="All editable inputs — blue text. Every formula in other tabs references cells here.")
apply_style(c, font=font_subtitle, alignment=align_center)

row = 4

# ── PROTOCOL FEES ──
style_section_header(ws_a, row, "PROTOCOL FEES", 6)
row += 1

fee_items = [
    ("Settlement Fee Rate", 0.025, "C", "On successful task completion"),
    ("Slash Fee Rate (protocol cut)", 0.10, "C", "Protocol's cut of slashed stake"),
    ("Dispute Fee Rate", 0.05, "C", "Fee charged on disputes"),
    ("Insurance Premium Cut", 0.01, "C", "Protocol cut of insurance premiums"),
]

# Map of assumption cell addresses (row -> cell ref)
assumptions = {}

for label, val, c_col, note in fee_items:
    ws_a.cell(row=row, column=2, value=label).font = font_label
    cell = ws_a.cell(row=row, column=3, value=val)
    apply_style(cell, font=font_input, fill=fill_input, number_format=NUM_PCT)
    ws_a.cell(row=row, column=4, value=note).font = font_subtitle
    assumptions[label] = f"C{row}"
    row += 1

row += 1

# ── FEE PHASING ──
style_section_header(ws_a, row, "FEE PHASING SCHEDULE", 6)
row += 1

phasing = [
    ("M1-3 Settlement Rate", 0.0, "Bootstrapping — zero fees"),
    ("M4-6 Settlement Rate", 0.01, "Introductory rate"),
    ("M7+ Settlement Rate", 0.025, "Full fee schedule"),
]

for label, val, note in phasing:
    ws_a.cell(row=row, column=2, value=label).font = font_label
    cell = ws_a.cell(row=row, column=3, value=val)
    apply_style(cell, font=font_input, fill=fill_input, number_format=NUM_PCT)
    ws_a.cell(row=row, column=4, value=note).font = font_subtitle
    assumptions[label] = f"C{row}"
    row += 1

row += 1

# ── SPOT TASKS ──
style_section_header(ws_a, row, "SPOT TASK ASSUMPTIONS", 6)
row += 1

spot_items = [
    ("Month 1 Spot Tasks", 20, NUM_INT, "Initial task volume"),
    ("Spot Task MoM Growth Rate", 0.30, NUM_PCT, "Monthly growth in task count"),
    ("Average Bounty (USD)", 1500, NUM_DOLLAR, "Starting average bounty"),
    ("Bounty MoM Growth Rate", 0.05, NUM_PCT, "Monthly increase in avg bounty"),
]

for label, val, fmt, note in spot_items:
    ws_a.cell(row=row, column=2, value=label).font = font_label
    cell = ws_a.cell(row=row, column=3, value=val)
    apply_style(cell, font=font_input, fill=fill_input, number_format=fmt)
    ws_a.cell(row=row, column=4, value=note).font = font_subtitle
    assumptions[label] = f"C{row}"
    row += 1

row += 1

# ── CONTINUOUS CONTRACTS ──
style_section_header(ws_a, row, "CONTINUOUS CONTRACT ASSUMPTIONS", 6)
row += 1

cc_items = [
    ("CC Launch Month", 3, NUM_INT, "Month CCs become available"),
    ("New CCs per Month", 2, NUM_INT, "New contracts added monthly"),
    ("CC Monthly Value (USD)", 5000, NUM_DOLLAR, "Revenue per active CC per month"),
    ("CC Retention Rate", 0.90, NUM_PCT, "Monthly retention"),
]

for label, val, fmt, note in cc_items:
    ws_a.cell(row=row, column=2, value=label).font = font_label
    cell = ws_a.cell(row=row, column=3, value=val)
    apply_style(cell, font=font_input, fill=fill_input, number_format=fmt)
    ws_a.cell(row=row, column=4, value=note).font = font_subtitle
    assumptions[label] = f"C{row}"
    row += 1

row += 1

# ── OUTCOME RATES ──
style_section_header(ws_a, row, "TASK OUTCOME RATES", 6)
row += 1

outcome_items = [
    ("Failure Rate", 0.20, NUM_PCT, "Tasks that fail / get slashed"),
    ("Stake/Bounty Ratio", 0.50, NUM_PCT, "Agent stake as % of bounty"),
    ("Avg Severity (BPS)", 3500, NUM_BPS, "Average slash severity in basis points"),
    ("Dispute Rate", 0.05, NUM_PCT, "% of tasks that go to dispute"),
    ("Cancellation Rate", 0.10, NUM_PCT, "% of tasks cancelled before completion"),
]

for label, val, fmt, note in outcome_items:
    ws_a.cell(row=row, column=2, value=label).font = font_label
    cell = ws_a.cell(row=row, column=3, value=val)
    apply_style(cell, font=font_input, fill=fill_input, number_format=fmt)
    ws_a.cell(row=row, column=4, value=note).font = font_subtitle
    assumptions[label] = f"C{row}"
    row += 1

row += 1

# ── INSURANCE ──
style_section_header(ws_a, row, "INSURANCE ASSUMPTIONS", 6)
row += 1

ins_items = [
    ("Insurance Launch Month", 4, NUM_INT, "Month insurance goes live"),
    ("Insurance Adoption Rate", 0.15, NUM_PCT, "% of tasks that buy insurance"),
    ("Insurance Premium Rate", 0.05, NUM_PCT, "Premium as % of bounty"),
]

for label, val, fmt, note in ins_items:
    ws_a.cell(row=row, column=2, value=label).font = font_label
    cell = ws_a.cell(row=row, column=3, value=val)
    apply_style(cell, font=font_input, fill=fill_input, number_format=fmt)
    ws_a.cell(row=row, column=4, value=note).font = font_subtitle
    assumptions[label] = f"C{row}"
    row += 1

row += 1

# ── SAAS DATA INTELLIGENCE ──
style_section_header(ws_a, row, "SAAS DATA INTELLIGENCE", 6)
row += 1

saas_items = [
    ("SaaS Launch Month", 6, NUM_INT, "Month SaaS product launches"),
    ("Initial SaaS Subscribers", 5, NUM_INT, "Subscribers at launch"),
    ("SaaS MoM Growth Rate", 0.20, NUM_PCT, "Monthly subscriber growth"),
    ("Tier 1 Basic Price (USD/mo)", 500, NUM_DOLLAR, "Basic tier monthly price"),
    ("Tier 1 Mix %", 0.60, NUM_PCT, "% of subscribers on Basic"),
    ("Tier 2 Pro Price (USD/mo)", 2000, NUM_DOLLAR, "Pro tier monthly price"),
    ("Tier 2 Mix %", 0.30, NUM_PCT, "% of subscribers on Pro"),
    ("Tier 3 Enterprise Price (USD/mo)", 8000, NUM_DOLLAR, "Enterprise tier monthly price"),
    ("Tier 3 Mix %", 0.10, NUM_PCT, "% of subscribers on Enterprise"),
]

for label, val, fmt, note in saas_items:
    ws_a.cell(row=row, column=2, value=label).font = font_label
    cell = ws_a.cell(row=row, column=3, value=val)
    apply_style(cell, font=font_input, fill=fill_input, number_format=fmt)
    ws_a.cell(row=row, column=4, value=note).font = font_subtitle
    assumptions[label] = f"C{row}"
    row += 1

# Calculated: Blended ARPU
ws_a.cell(row=row, column=2, value="Blended ARPU (USD/mo)").font = font_formula_bold
arpu_formula = (
    f"={assumptions['Tier 1 Basic Price (USD/mo)']}*{assumptions['Tier 1 Mix %']}"
    f"+{assumptions['Tier 2 Pro Price (USD/mo)']}*{assumptions['Tier 2 Mix %']}"
    f"+{assumptions['Tier 3 Enterprise Price (USD/mo)']}*{assumptions['Tier 3 Mix %']}"
)
cell = ws_a.cell(row=row, column=3, value=arpu_formula)
apply_style(cell, font=font_formula_bold, number_format=NUM_DOLLAR)
ws_a.cell(row=row, column=4, value="Calculated: weighted avg of tiers").font = font_subtitle
assumptions["Blended ARPU"] = f"C{row}"
row += 2

# ── COSTS ──
style_section_header(ws_a, row, "COST ASSUMPTIONS", 6)
row += 1

cost_items = [
    ("One-Time: Audit Cost", 15000, NUM_DOLLAR, "Smart contract audit (M1)"),
    ("One-Time: Legal Cost", 5000, NUM_DOLLAR, "Legal setup (M1)"),
    ("One-Time: Seed Tasks", 5000, NUM_DOLLAR, "Seed task bounties (M1)"),
    ("Monthly Infra Cost", 200, NUM_DOLLAR, "Protocol infrastructure (M1 start)"),
    ("Infra MoM Growth Rate", 0.10, NUM_PCT, "Monthly infra cost increase"),
    ("Gas Cost per Task", 0.50, NUM_DOLLAR_2, "On-chain gas per task"),
    ("SaaS Infra Cost (USD/mo)", 500, NUM_DOLLAR, "SaaS platform costs (from launch)"),
    ("SaaS Infra Growth Rate", 0.08, NUM_PCT, "Monthly SaaS infra growth"),
    ("Marketing Launch Month", 3, NUM_INT, "Month marketing starts"),
    ("Marketing Start Cost (USD/mo)", 1000, NUM_DOLLAR, "Initial marketing spend"),
    ("Marketing MoM Growth Rate", 0.15, NUM_PCT, "Monthly marketing cost growth"),
]

for label, val, fmt, note in cost_items:
    ws_a.cell(row=row, column=2, value=label).font = font_label
    cell = ws_a.cell(row=row, column=3, value=val)
    apply_style(cell, font=font_input, fill=fill_input, number_format=fmt)
    ws_a.cell(row=row, column=4, value=note).font = font_subtitle
    assumptions[label] = f"C{row}"
    row += 1


# Print assumptions map for debugging
print("=== ASSUMPTIONS CELL MAP ===")
for k, v in assumptions.items():
    print(f"  {k}: {v}")

# Helper to get assumption ref
def a(name):
    return f"'Assumptions'!{assumptions[name]}"


# ═══════════════════════════════════════════════════
# TAB 2: P&L (24-Month)
# ═══════════════════════════════════════════════════

ws_pl = wb.create_sheet("P&L")
ws_pl.sheet_properties.tabColor = NAVY
ws_pl.sheet_view.showGridLines = False

# Columns: A=row#/spacer, B=label, C=M1...Z=M24, AA=Y1 Total, AB=Y2 Total
# C=3 is M1, so M24 = col 26 (Z), Y1=AA(27), Y2=AB(28)
PL_LABEL_COL = 2
PL_M1_COL = 3  # Column C = Month 1
PL_LAST_COL = 28  # AB = Y2 Total
PL_Y1_COL = 27  # AA
PL_Y2_COL = 28  # AB

ws_pl.column_dimensions['A'].width = 3
ws_pl.column_dimensions['B'].width = 32

for c in range(PL_M1_COL, PL_LAST_COL + 1):
    ws_pl.column_dimensions[col(c)].width = 14

# Row 1: Title
ws_pl.merge_cells(f'A1:{col(PL_LAST_COL)}1')
c_title = ws_pl.cell(row=1, column=1, value="THE ARENA PROTOCOL — 24-MONTH P&L MODEL")
apply_style(c_title, font=font_title, fill=fill_navy, alignment=align_center)
for cc in range(2, PL_LAST_COL + 1):
    apply_style(ws_pl.cell(row=1, column=cc), fill=fill_navy)

# Row 2: subtitle
ws_pl.merge_cells(f'A2:{col(PL_LAST_COL)}2')
c_sub = ws_pl.cell(row=2, column=1, value="All figures in USD. Blue = hardcoded input, Black = formula referencing Assumptions tab.")
apply_style(c_sub, font=font_subtitle, alignment=align_center)

# Row 3: Month numbers header
ws_pl.cell(row=3, column=PL_LABEL_COL, value="").font = font_header
for m in range(1, 25):
    c_cell = ws_pl.cell(row=3, column=PL_M1_COL + m - 1, value=f"M{m}")
    apply_style(c_cell, font=font_header, fill=fill_navy, alignment=align_center)

ws_pl.cell(row=3, column=PL_Y1_COL, value="Y1 Total")
apply_style(ws_pl.cell(row=3, column=PL_Y1_COL), font=font_header, fill=fill_light_navy, alignment=align_center)
ws_pl.cell(row=3, column=PL_Y2_COL, value="Y2 Total")
apply_style(ws_pl.cell(row=3, column=PL_Y2_COL), font=font_header, fill=fill_light_navy, alignment=align_center)

# Apply navy fill to row 3 col A, B
apply_style(ws_pl.cell(row=3, column=1), fill=fill_navy)
apply_style(ws_pl.cell(row=3, column=2), fill=fill_navy)

# Freeze panes: row 4, column C
ws_pl.freeze_panes = "C4"

# ── P&L ROW BUILDER ──
# We'll track which row each metric is on
pl_rows = {}
current_row = 4


def pl_section(label):
    global current_row
    current_row += 1
    style_section_header(ws_pl, current_row, label, PL_LAST_COL)
    current_row += 1


def pl_row(label, formulas_by_month, fmt=NUM_DOLLAR, is_total=False, is_subtotal=False):
    """
    Add a P&L row.
    formulas_by_month: dict of {month_num: formula_string} for months 1-24
    If a month is not in the dict, cell is left blank.
    """
    global current_row
    r = current_row
    pl_rows[label] = r

    ws_pl.cell(row=r, column=PL_LABEL_COL, value=label)

    if is_total:
        ws_pl.cell(row=r, column=PL_LABEL_COL).font = font_total
    elif is_subtotal:
        ws_pl.cell(row=r, column=PL_LABEL_COL).font = font_formula_bold
    else:
        ws_pl.cell(row=r, column=PL_LABEL_COL).font = font_label

    for m in range(1, 25):
        c_idx = PL_M1_COL + m - 1
        if m in formulas_by_month:
            cell = ws_pl.cell(row=r, column=c_idx, value=formulas_by_month[m])
            f = font_total if is_total else (font_formula_bold if is_subtotal else font_formula)
            apply_style(cell, font=f, number_format=fmt, alignment=align_right)

    # Y1 total (sum M1-M12)
    y1_start = col(PL_M1_COL)
    y1_end = col(PL_M1_COL + 11)
    y1_formula = f"=SUM({y1_start}{r}:{y1_end}{r})"
    cell_y1 = ws_pl.cell(row=r, column=PL_Y1_COL, value=y1_formula)
    apply_style(cell_y1, font=font_formula_bold, number_format=fmt, alignment=align_right, fill=fill_light_gray)

    # Y2 total (sum M13-M24)
    y2_start = col(PL_M1_COL + 12)
    y2_end = col(PL_M1_COL + 23)
    y2_formula = f"=SUM({y2_start}{r}:{y2_end}{r})"
    cell_y2 = ws_pl.cell(row=r, column=PL_Y2_COL, value=y2_formula)
    apply_style(cell_y2, font=font_formula_bold, number_format=fmt, alignment=align_right, fill=fill_light_gray)

    if is_total:
        style_total_row(ws_pl, r, PL_LAST_COL)

    current_row += 1
    return r


def pl_row_no_yearly(label, formulas_by_month, fmt=NUM_PCT, is_total=False):
    """P&L row without Y1/Y2 sums (for rates/ratios)."""
    global current_row
    r = current_row
    pl_rows[label] = r

    ws_pl.cell(row=r, column=PL_LABEL_COL, value=label)
    ws_pl.cell(row=r, column=PL_LABEL_COL).font = font_total if is_total else font_label

    for m in range(1, 25):
        c_idx = PL_M1_COL + m - 1
        if m in formulas_by_month:
            cell = ws_pl.cell(row=r, column=c_idx, value=formulas_by_month[m])
            f = font_total if is_total else font_formula
            apply_style(cell, font=f, number_format=fmt, alignment=align_right)

    current_row += 1
    return r


def pl_blank():
    global current_row
    current_row += 1


def mcol(month):
    """Get column letter for a given month (1-24) in P&L."""
    return col(PL_M1_COL + month - 1)


def mcell(row_label, month):
    """Get cell reference for a P&L metric at a given month."""
    r = pl_rows[row_label]
    c = mcol(month)
    return f"{c}{r}"


# ═══════════════════════════════
# VOLUME SECTION
# ═══════════════════════════════
pl_section("VOLUME")

# Spot Tasks: M1 = from assumptions, M2+ = prev * (1 + growth)
spot_formulas = {}
for m in range(1, 25):
    if m == 1:
        spot_formulas[m] = f"={a('Month 1 Spot Tasks')}"
    else:
        prev_col = mcol(m - 1)
        spot_formulas[m] = f"=ROUND({prev_col}{{ROW}}*(1+{a('Spot Task MoM Growth Rate')}),0)"
# We need to resolve {{ROW}} after we know the row
# Instead, let's build formulas knowing the row will be current_row
spot_row = current_row
for m in range(1, 25):
    if m == 1:
        spot_formulas[m] = f"={a('Month 1 Spot Tasks')}"
    else:
        prev_col = mcol(m - 1)
        spot_formulas[m] = f"=ROUND({prev_col}{spot_row}*(1+{a('Spot Task MoM Growth Rate')}),0)"
pl_row("Spot Tasks", spot_formulas, fmt=NUM_INT)

# Avg Bounty: M1 = assumption, M2+ = prev * (1 + bounty growth)
bounty_row = current_row
bounty_formulas = {}
for m in range(1, 25):
    if m == 1:
        bounty_formulas[m] = f"={a('Average Bounty (USD)')}"
    else:
        prev_col = mcol(m - 1)
        bounty_formulas[m] = f"={prev_col}{bounty_row}*(1+{a('Bounty MoM Growth Rate')})"
pl_row("Avg Bounty (USD)", bounty_formulas, fmt=NUM_DOLLAR)

# Spot GMV: tasks * bounty
spot_gmv_row = current_row
spot_gmv_formulas = {}
for m in range(1, 25):
    spot_gmv_formulas[m] = f"={mcol(m)}{pl_rows['Spot Tasks']}*{mcol(m)}{pl_rows['Avg Bounty (USD)']}"
pl_row("Spot GMV", spot_gmv_formulas, fmt=NUM_DOLLAR)

# Active CCs: 0 before launch, accumulate with retention
cc_row = current_row
cc_formulas = {}
for m in range(1, 25):
    launch = assumptions['CC Launch Month']
    # IF(month < launch, 0, prev * retention + new)
    if m == 1:
        cc_formulas[m] = f"=IF(1<{a('CC Launch Month')},0,{a('New CCs per Month')})"
    else:
        prev_col = mcol(m - 1)
        cc_formulas[m] = (
            f"=IF({m}<{a('CC Launch Month')},0,"
            f"ROUND({prev_col}{cc_row}*{a('CC Retention Rate')}+{a('New CCs per Month')},0))"
        )
pl_row("Active CCs", cc_formulas, fmt=NUM_INT)

# CC GMV: active CCs * monthly value
cc_gmv_row = current_row
cc_gmv_formulas = {}
for m in range(1, 25):
    cc_gmv_formulas[m] = f"={mcol(m)}{pl_rows['Active CCs']}*{a('CC Monthly Value (USD)')}"
pl_row("CC GMV", cc_gmv_formulas, fmt=NUM_DOLLAR)

# Total Tasks = Spot Tasks (CCs are continuous, not per-task; we count spot tasks + CC equivalents)
# For total tasks, CC tasks = active CCs (each CC = ~1 task-equivalent per month for gas/volume purposes)
total_tasks_row = current_row
tt_formulas = {}
for m in range(1, 25):
    tt_formulas[m] = f"={mcol(m)}{pl_rows['Spot Tasks']}+{mcol(m)}{pl_rows['Active CCs']}"
pl_row("Total Tasks", tt_formulas, fmt=NUM_INT, is_subtotal=True)

# TOTAL GMV
total_gmv_row = current_row
tg_formulas = {}
for m in range(1, 25):
    tg_formulas[m] = f"={mcol(m)}{pl_rows['Spot GMV']}+{mcol(m)}{pl_rows['CC GMV']}"
pl_row("TOTAL GMV", tg_formulas, fmt=NUM_DOLLAR, is_total=True)

pl_blank()

# ═══════════════════════════════
# REVENUE SECTION
# ═══════════════════════════════
pl_section("REVENUE")

# Effective Settlement Rate (uses phasing)
eff_rate_row = current_row
eff_rate_formulas = {}
for m in range(1, 25):
    eff_rate_formulas[m] = (
        f"=IF({m}<=3,{a('M1-3 Settlement Rate')},"
        f"IF({m}<=6,{a('M4-6 Settlement Rate')},"
        f"{a('M7+ Settlement Rate')}))"
    )
pl_row_no_yearly("Effective Settlement Rate", eff_rate_formulas, fmt=NUM_PCT)

# Settlement Fees: GMV * (1 - cancel) * (1 - fail) * effective rate
settle_row = current_row
settle_formulas = {}
for m in range(1, 25):
    settle_formulas[m] = (
        f"={mcol(m)}{pl_rows['TOTAL GMV']}"
        f"*(1-{a('Cancellation Rate')})"
        f"*(1-{a('Failure Rate')})"
        f"*{mcol(m)}{pl_rows['Effective Settlement Rate']}"
    )
pl_row("Settlement Fees", settle_formulas, fmt=NUM_DOLLAR)

# Failed Task Count: total tasks * (1-cancel) * failure rate
failed_row = current_row
failed_formulas = {}
for m in range(1, 25):
    failed_formulas[m] = (
        f"=ROUND({mcol(m)}{pl_rows['Total Tasks']}"
        f"*(1-{a('Cancellation Rate')})"
        f"*{a('Failure Rate')},0)"
    )
pl_row("Failed Tasks", failed_formulas, fmt=NUM_INT)

# Slash Revenue: failed * avg bounty * stake ratio * severity/10000 * slash fee rate
slash_row = current_row
slash_formulas = {}
for m in range(1, 25):
    slash_formulas[m] = (
        f"={mcol(m)}{pl_rows['Failed Tasks']}"
        f"*{mcol(m)}{pl_rows['Avg Bounty (USD)']}"
        f"*{a('Stake/Bounty Ratio')}"
        f"*{a('Avg Severity (BPS)')}/10000"
        f"*{a('Slash Fee Rate (protocol cut)')}"
    )
pl_row("Slash Revenue", slash_formulas, fmt=NUM_DOLLAR)

# Dispute Fees: total tasks * (1-cancel) * dispute rate * avg bounty * dispute fee rate
dispute_row = current_row
dispute_formulas = {}
for m in range(1, 25):
    dispute_formulas[m] = (
        f"={mcol(m)}{pl_rows['Total Tasks']}"
        f"*(1-{a('Cancellation Rate')})"
        f"*{a('Dispute Rate')}"
        f"*{mcol(m)}{pl_rows['Avg Bounty (USD)']}"
        f"*{a('Dispute Fee Rate')}"
    )
pl_row("Dispute Fees", dispute_formulas, fmt=NUM_DOLLAR)

# Insurance Premiums: IF month >= launch, total tasks * (1-cancel) * adoption * avg bounty * premium rate * protocol cut
ins_row = current_row
ins_formulas = {}
for m in range(1, 25):
    ins_formulas[m] = (
        f"=IF({m}<{a('Insurance Launch Month')},0,"
        f"{mcol(m)}{pl_rows['Total Tasks']}"
        f"*(1-{a('Cancellation Rate')})"
        f"*{a('Insurance Adoption Rate')}"
        f"*{mcol(m)}{pl_rows['Avg Bounty (USD)']}"
        f"*{a('Insurance Premium Rate')}"
        f"*{a('Insurance Premium Cut')})"  # protocol's cut
    )
pl_row("Insurance Premiums", ins_formulas, fmt=NUM_DOLLAR)

# SaaS Subscribers: compound growth from launch
saas_subs_row = current_row
saas_subs_formulas = {}
for m in range(1, 25):
    if m == 1:
        saas_subs_formulas[m] = (
            f"=IF(1<{a('SaaS Launch Month')},0,{a('Initial SaaS Subscribers')})"
        )
    else:
        prev = mcol(m - 1)
        saas_subs_formulas[m] = (
            f"=IF({m}<{a('SaaS Launch Month')},0,"
            f"IF({m}={a('SaaS Launch Month')},{a('Initial SaaS Subscribers')},"
            f"ROUND({prev}{saas_subs_row}*(1+{a('SaaS MoM Growth Rate')}),0)))"
        )
pl_row("SaaS Subscribers", saas_subs_formulas, fmt=NUM_INT)

# SaaS Revenue: subscribers * blended ARPU
saas_rev_row = current_row
saas_rev_formulas = {}
for m in range(1, 25):
    saas_rev_formulas[m] = f"={mcol(m)}{pl_rows['SaaS Subscribers']}*{a('Blended ARPU')}"
pl_row("SaaS Revenue", saas_rev_formulas, fmt=NUM_DOLLAR)

# TOTAL REVENUE
total_rev_row = current_row
tr_formulas = {}
for m in range(1, 25):
    tr_formulas[m] = (
        f"={mcol(m)}{pl_rows['Settlement Fees']}"
        f"+{mcol(m)}{pl_rows['Slash Revenue']}"
        f"+{mcol(m)}{pl_rows['Dispute Fees']}"
        f"+{mcol(m)}{pl_rows['Insurance Premiums']}"
        f"+{mcol(m)}{pl_rows['SaaS Revenue']}"
    )
pl_row("TOTAL REVENUE", tr_formulas, fmt=NUM_DOLLAR, is_total=True)

pl_blank()

# ═══════════════════════════════
# COSTS SECTION
# ═══════════════════════════════
pl_section("COSTS")

# One-Time Costs (M1 only)
onetime_row = current_row
onetime_formulas = {}
onetime_formulas[1] = (
    f"={a('One-Time: Audit Cost')}+{a('One-Time: Legal Cost')}+{a('One-Time: Seed Tasks')}"
)
for m in range(2, 25):
    onetime_formulas[m] = "=0"
pl_row("One-Time Costs", onetime_formulas, fmt=NUM_DOLLAR)

# Protocol Infra: base * (1+growth)^(m-1)
infra_row = current_row
infra_formulas = {}
for m in range(1, 25):
    infra_formulas[m] = (
        f"={a('Monthly Infra Cost')}*(1+{a('Infra MoM Growth Rate')})^({m}-1)"
    )
pl_row("Protocol Infrastructure", infra_formulas, fmt=NUM_DOLLAR)

# Gas Costs: total tasks * gas/task
gas_row = current_row
gas_formulas = {}
for m in range(1, 25):
    gas_formulas[m] = f"={mcol(m)}{pl_rows['Total Tasks']}*{a('Gas Cost per Task')}"
pl_row("Gas Costs", gas_formulas, fmt=NUM_DOLLAR)

# SaaS Infra: IF month >= saas launch, base * (1+growth)^(months since launch)
saas_infra_row = current_row
saas_infra_formulas = {}
for m in range(1, 25):
    saas_infra_formulas[m] = (
        f"=IF({m}<{a('SaaS Launch Month')},0,"
        f"{a('SaaS Infra Cost (USD/mo)')}*(1+{a('SaaS Infra Growth Rate')})^({m}-{a('SaaS Launch Month')}))"
    )
pl_row("SaaS Infrastructure", saas_infra_formulas, fmt=NUM_DOLLAR)

# Marketing: IF month >= marketing launch, base * (1+growth)^(months since launch)
mkt_row = current_row
mkt_formulas = {}
for m in range(1, 25):
    mkt_formulas[m] = (
        f"=IF({m}<{a('Marketing Launch Month')},0,"
        f"{a('Marketing Start Cost (USD/mo)')}*(1+{a('Marketing MoM Growth Rate')})^({m}-{a('Marketing Launch Month')}))"
    )
pl_row("Marketing", mkt_formulas, fmt=NUM_DOLLAR)

# TOTAL COSTS
total_cost_row = current_row
tc_formulas = {}
for m in range(1, 25):
    tc_formulas[m] = (
        f"={mcol(m)}{pl_rows['One-Time Costs']}"
        f"+{mcol(m)}{pl_rows['Protocol Infrastructure']}"
        f"+{mcol(m)}{pl_rows['Gas Costs']}"
        f"+{mcol(m)}{pl_rows['SaaS Infrastructure']}"
        f"+{mcol(m)}{pl_rows['Marketing']}"
    )
pl_row("TOTAL COSTS", tc_formulas, fmt=NUM_DOLLAR, is_total=True)

pl_blank()

# ═══════════════════════════════
# PROFITABILITY SECTION
# ═══════════════════════════════
pl_section("PROFITABILITY")

# Net Income = Revenue - Costs
ni_row = current_row
ni_formulas = {}
for m in range(1, 25):
    ni_formulas[m] = (
        f"={mcol(m)}{pl_rows['TOTAL REVENUE']}-{mcol(m)}{pl_rows['TOTAL COSTS']}"
    )
pl_row("NET INCOME", ni_formulas, fmt=NUM_DOLLAR, is_total=True)

# Cumulative P&L
cum_row = current_row
cum_formulas = {}
for m in range(1, 25):
    if m == 1:
        cum_formulas[m] = f"={mcol(1)}{pl_rows['NET INCOME']}"
    else:
        cum_formulas[m] = f"={mcol(m-1)}{cum_row}+{mcol(m)}{pl_rows['NET INCOME']}"
pl_row("Cumulative P&L", cum_formulas, fmt=NUM_DOLLAR, is_subtotal=True)

# Net Margin = Net Income / Revenue (avoid div/0)
margin_row = current_row
margin_formulas = {}
for m in range(1, 25):
    margin_formulas[m] = (
        f"=IFERROR({mcol(m)}{pl_rows['NET INCOME']}/{mcol(m)}{pl_rows['TOTAL REVENUE']},0)"
    )
pl_row_no_yearly("Net Margin", margin_formulas, fmt=NUM_PCT)

# Take Rate = Revenue / GMV
take_row = current_row
take_formulas = {}
for m in range(1, 25):
    take_formulas[m] = (
        f"=IFERROR({mcol(m)}{pl_rows['TOTAL REVENUE']}/{mcol(m)}{pl_rows['TOTAL GMV']},0)"
    )
pl_row_no_yearly("Take Rate", take_formulas, fmt=NUM_PCT)

# Revenue per Task
rpt_row = current_row
rpt_formulas = {}
for m in range(1, 25):
    rpt_formulas[m] = (
        f"=IFERROR({mcol(m)}{pl_rows['TOTAL REVENUE']}/{mcol(m)}{pl_rows['Total Tasks']},0)"
    )
pl_row_no_yearly("Revenue per Task", rpt_formulas, fmt=NUM_DOLLAR)

pl_blank()

# ═══════════════════════════════
# KEY METRICS
# ═══════════════════════════════
pl_section("KEY METRICS")

# Cumulative Tasks
cum_tasks_row = current_row
ct_formulas = {}
for m in range(1, 25):
    if m == 1:
        ct_formulas[m] = f"={mcol(1)}{pl_rows['Total Tasks']}"
    else:
        ct_formulas[m] = f"={mcol(m-1)}{cum_tasks_row}+{mcol(m)}{pl_rows['Total Tasks']}"
pl_row("Cumulative Tasks", ct_formulas, fmt=NUM_INT, is_subtotal=True)

# Cumulative GMV
cum_gmv_row = current_row
cg_formulas = {}
for m in range(1, 25):
    if m == 1:
        cg_formulas[m] = f"={mcol(1)}{pl_rows['TOTAL GMV']}"
    else:
        cg_formulas[m] = f"={mcol(m-1)}{cum_gmv_row}+{mcol(m)}{pl_rows['TOTAL GMV']}"
pl_row("Cumulative GMV", cg_formulas, fmt=NUM_DOLLAR, is_subtotal=True)

# SaaS ARR (monthly revenue * 12)
saas_arr_row = current_row
sa_formulas = {}
for m in range(1, 25):
    sa_formulas[m] = f"={mcol(m)}{pl_rows['SaaS Revenue']}*12"
pl_row("SaaS ARR", sa_formulas, fmt=NUM_DOLLAR)

# Total ARR (total monthly revenue * 12)
total_arr_row = current_row
ta_formulas = {}
for m in range(1, 25):
    ta_formulas[m] = f"={mcol(m)}{pl_rows['TOTAL REVENUE']}*12"
pl_row("Total ARR", ta_formulas, fmt=NUM_DOLLAR, is_subtotal=True)


# ═══════════════════════════════════════════════════
# TAB 3: UNIT ECONOMICS
# ═══════════════════════════════════════════════════

ws_ue = wb.create_sheet("Unit Economics")
ws_ue.sheet_properties.tabColor = "2C3E6B"
ws_ue.sheet_view.showGridLines = False

ws_ue.column_dimensions['A'].width = 3
ws_ue.column_dimensions['B'].width = 38
ws_ue.column_dimensions['C'].width = 22
ws_ue.column_dimensions['D'].width = 22
ws_ue.column_dimensions['E'].width = 22

# Title
ws_ue.merge_cells('A1:E1')
c = ws_ue.cell(row=1, column=1, value="UNIT ECONOMICS — PER-TASK ANALYSIS")
apply_style(c, font=font_title, fill=fill_navy, alignment=align_center)
for cc in range(2, 6):
    apply_style(ws_ue.cell(row=1, column=cc), fill=fill_navy)

r = 3

# Inputs section
style_section_header(ws_ue, r, "INPUTS", 5)
r += 1

ws_ue.cell(row=r, column=2, value="Example Bounty").font = font_label
ws_ue.cell(row=r, column=3, value=2500)
apply_style(ws_ue.cell(row=r, column=3), font=font_input, fill=fill_input, number_format=NUM_DOLLAR)
ue_bounty = f"C{r}"
r += 1

ws_ue.cell(row=r, column=2, value="Agent Price (winning bid)").font = font_label
ws_ue.cell(row=r, column=3, value=2200)
apply_style(ws_ue.cell(row=r, column=3), font=font_input, fill=fill_input, number_format=NUM_DOLLAR)
ue_price = f"C{r}"
r += 1

ws_ue.cell(row=r, column=2, value="Settlement Fee Rate").font = font_label
ws_ue.cell(row=r, column=3, value=f"={a('Settlement Fee Rate')}")
apply_style(ws_ue.cell(row=r, column=3), font=font_formula, number_format=NUM_PCT)
ue_settle = f"C{r}"
r += 1

ws_ue.cell(row=r, column=2, value="Stake/Bounty Ratio").font = font_label
ws_ue.cell(row=r, column=3, value=f"={a('Stake/Bounty Ratio')}")
apply_style(ws_ue.cell(row=r, column=3), font=font_formula, number_format=NUM_PCT)
ue_stake_ratio = f"C{r}"
r += 1

ws_ue.cell(row=r, column=2, value="Avg Severity (BPS)").font = font_label
ws_ue.cell(row=r, column=3, value=f"={a('Avg Severity (BPS)')}")
apply_style(ws_ue.cell(row=r, column=3), font=font_formula, number_format=NUM_BPS)
ue_severity = f"C{r}"
r += 1

ws_ue.cell(row=r, column=2, value="Slash Fee Rate").font = font_label
ws_ue.cell(row=r, column=3, value=f"={a('Slash Fee Rate (protocol cut)')}")
apply_style(ws_ue.cell(row=r, column=3), font=font_formula, number_format=NUM_PCT)
ue_slash = f"C{r}"
r += 1

ws_ue.cell(row=r, column=2, value="Gas Cost per Task").font = font_label
ws_ue.cell(row=r, column=3, value=f"={a('Gas Cost per Task')}")
apply_style(ws_ue.cell(row=r, column=3), font=font_formula, number_format=NUM_DOLLAR_2)
ue_gas = f"C{r}"
r += 1

ws_ue.cell(row=r, column=2, value="Failure Rate").font = font_label
ws_ue.cell(row=r, column=3, value=f"={a('Failure Rate')}")
apply_style(ws_ue.cell(row=r, column=3), font=font_formula, number_format=NUM_PCT)
ue_fail = f"C{r}"
r += 2

# Success Path
style_section_header(ws_ue, r, "SUCCESS PATH (Task Completed)", 5)
r += 1

ws_ue.cell(row=r, column=2, value="Settlement Fee").font = font_label
ws_ue.cell(row=r, column=3, value=f"={ue_price}*{ue_settle}")
apply_style(ws_ue.cell(row=r, column=3), font=font_formula, number_format=NUM_DOLLAR_2)
ue_settle_fee = f"C{r}"
r += 1

ws_ue.cell(row=r, column=2, value="Gas Cost").font = font_label
ws_ue.cell(row=r, column=3, value=f"=-{ue_gas}")
apply_style(ws_ue.cell(row=r, column=3), font=font_formula, number_format=NUM_DOLLAR_2)
ue_gas_cost_s = f"C{r}"
r += 1

ws_ue.cell(row=r, column=2, value="Net Revenue (Success)").font = font_formula_bold
ws_ue.cell(row=r, column=3, value=f"={ue_settle_fee}+{ue_gas_cost_s}")
apply_style(ws_ue.cell(row=r, column=3), font=font_formula_bold, number_format=NUM_DOLLAR_2, border=border_bottom_thick)
ue_net_success = f"C{r}"
r += 2

# Failure Path
style_section_header(ws_ue, r, "FAILURE PATH (Task Slashed)", 5)
r += 1

ws_ue.cell(row=r, column=2, value="Stake Amount").font = font_label
ws_ue.cell(row=r, column=3, value=f"={ue_bounty}*{ue_stake_ratio}")
apply_style(ws_ue.cell(row=r, column=3), font=font_formula, number_format=NUM_DOLLAR_2)
ue_stake_amt = f"C{r}"
r += 1

ws_ue.cell(row=r, column=2, value="Slash Amount (at severity)").font = font_label
ws_ue.cell(row=r, column=3, value=f"={ue_stake_amt}*{ue_severity}/10000")
apply_style(ws_ue.cell(row=r, column=3), font=font_formula, number_format=NUM_DOLLAR_2)
ue_slash_amt = f"C{r}"
r += 1

ws_ue.cell(row=r, column=2, value="Protocol Slash Revenue").font = font_label
ws_ue.cell(row=r, column=3, value=f"={ue_slash_amt}*{ue_slash}")
apply_style(ws_ue.cell(row=r, column=3), font=font_formula, number_format=NUM_DOLLAR_2)
ue_proto_slash = f"C{r}"
r += 1

ws_ue.cell(row=r, column=2, value="Gas Cost").font = font_label
ws_ue.cell(row=r, column=3, value=f"=-{ue_gas}")
apply_style(ws_ue.cell(row=r, column=3), font=font_formula, number_format=NUM_DOLLAR_2)
ue_gas_cost_f = f"C{r}"
r += 1

ws_ue.cell(row=r, column=2, value="Net Revenue (Failure)").font = font_formula_bold
ws_ue.cell(row=r, column=3, value=f"={ue_proto_slash}+{ue_gas_cost_f}")
apply_style(ws_ue.cell(row=r, column=3), font=font_formula_bold, number_format=NUM_DOLLAR_2, border=border_bottom_thick)
ue_net_failure = f"C{r}"
r += 2

# Blended
style_section_header(ws_ue, r, "BLENDED ECONOMICS (80% Success + 20% Failure)", 5)
r += 1

ws_ue.cell(row=r, column=2, value="Blended Revenue per Task").font = font_formula_bold
ws_ue.cell(row=r, column=3, value=f"=(1-{ue_fail})*{ue_net_success}+{ue_fail}*{ue_net_failure}")
apply_style(ws_ue.cell(row=r, column=3), font=font_formula_bold, number_format=NUM_DOLLAR_2)
ue_blended = f"C{r}"
r += 1

ws_ue.cell(row=r, column=2, value="Blended Margin per Task").font = font_formula_bold
ws_ue.cell(row=r, column=3, value=f"=IFERROR({ue_blended}/{ue_bounty},0)")
apply_style(ws_ue.cell(row=r, column=3), font=font_formula_bold, number_format=NUM_PCT, border=border_top_double)


# ═══════════════════════════════════════════════════
# TAB 4: SAAS INTELLIGENCE
# ═══════════════════════════════════════════════════

ws_saas = wb.create_sheet("SaaS Intelligence")
ws_saas.sheet_properties.tabColor = "2C3E6B"
ws_saas.sheet_view.showGridLines = False

ws_saas.column_dimensions['A'].width = 3
ws_saas.column_dimensions['B'].width = 35
ws_saas.column_dimensions['C'].width = 18
ws_saas.column_dimensions['D'].width = 18
ws_saas.column_dimensions['E'].width = 18
ws_saas.column_dimensions['F'].width = 18
ws_saas.column_dimensions['G'].width = 18

# Title
ws_saas.merge_cells('A1:G1')
c = ws_saas.cell(row=1, column=1, value="SAAS DATA INTELLIGENCE — AGENT RELIABILITY INDEX")
apply_style(c, font=font_title, fill=fill_navy, alignment=align_center)
for cc in range(2, 8):
    apply_style(ws_saas.cell(row=1, column=cc), fill=fill_navy)

r = 3
style_section_header(ws_saas, r, "PRODUCT DESCRIPTION", 7)
r += 1
ws_saas.merge_cells(f'B{r}:G{r}')
ws_saas.cell(row=r, column=2, value="The Agent Reliability Index (ARI) is a data product built on on-chain AI agent performance history.").font = font_label
r += 1
ws_saas.merge_cells(f'B{r}:G{r}')
ws_saas.cell(row=r, column=2, value="It sells AI agent performance scoring, task outcome predictions, and risk analytics derived from Arena protocol execution data.").font = font_label
r += 2

# Pricing Tiers
style_section_header(ws_saas, r, "PRICING TIERS", 7)
r += 1

# Header
for i, h in enumerate(["Tier", "Price/mo", "Mix %", "Features"], 2):
    c = ws_saas.cell(row=r, column=i, value=h)
    apply_style(c, font=font_white_bold, fill=fill_light_navy, alignment=align_center)
r += 1

tiers = [
    ("Basic", f"={a('Tier 1 Basic Price (USD/mo)')}", f"={a('Tier 1 Mix %')}", "Agent scores, basic reliability ratings, weekly reports"),
    ("Pro", f"={a('Tier 2 Pro Price (USD/mo)')}", f"={a('Tier 2 Mix %')}", "Real-time alerts, task outcome predictions, API access, custom dashboards"),
    ("Enterprise", f"={a('Tier 3 Enterprise Price (USD/mo)')}", f"={a('Tier 3 Mix %')}", "White-label data, custom models, SLA guarantees, dedicated support"),
]

for tier_name, price, mix, features in tiers:
    ws_saas.cell(row=r, column=2, value=tier_name).font = font_formula_bold
    ws_saas.cell(row=r, column=3, value=price)
    apply_style(ws_saas.cell(row=r, column=3), font=font_formula, number_format=NUM_DOLLAR)
    ws_saas.cell(row=r, column=4, value=mix)
    apply_style(ws_saas.cell(row=r, column=4), font=font_formula, number_format=NUM_PCT)
    ws_saas.cell(row=r, column=5, value=features).font = font_label
    r += 1

ws_saas.cell(row=r, column=2, value="Blended ARPU").font = font_total
ws_saas.cell(row=r, column=3, value=f"={a('Blended ARPU')}")
apply_style(ws_saas.cell(row=r, column=3), font=font_total, number_format=NUM_DOLLAR, border=border_top_double)
r += 2

# SaaS Projections table
style_section_header(ws_saas, r, "SAAS PROJECTIONS", 7)
r += 1

# Column headers: Metric, M6, M9, M12, M18, M24
saas_months = [6, 9, 12, 18, 24]
headers_saas = ["Metric"] + [f"M{m}" for m in saas_months]
for i, h in enumerate(headers_saas):
    c = ws_saas.cell(row=r, column=2 + i, value=h)
    apply_style(c, font=font_white_bold, fill=fill_light_navy, alignment=align_center)
r += 1

# Rows pulling from P&L
saas_metrics = [
    ("Subscribers", "SaaS Subscribers", NUM_INT),
    ("Monthly Revenue", "SaaS Revenue", NUM_DOLLAR),
    ("SaaS ARR", "SaaS ARR", NUM_DOLLAR),
    ("SaaS Infra Cost", "SaaS Infrastructure", NUM_DOLLAR),
]

for label, pl_label, fmt in saas_metrics:
    ws_saas.cell(row=r, column=2, value=label).font = font_label
    for i, m in enumerate(saas_months):
        pl_r = pl_rows[pl_label]
        m_col = mcol(m)
        ws_saas.cell(row=r, column=3 + i, value=f"='P&L'!{m_col}{pl_r}")
        apply_style(ws_saas.cell(row=r, column=3 + i), font=font_formula, number_format=fmt)
    r += 1

# Gross Profit = SaaS Revenue - SaaS Infra
ws_saas.cell(row=r, column=2, value="Gross Profit").font = font_total
for i, m in enumerate(saas_months):
    rev_r = pl_rows['SaaS Revenue']
    cost_r = pl_rows['SaaS Infrastructure']
    m_c = mcol(m)
    ws_saas.cell(row=r, column=3 + i, value=f"='P&L'!{m_c}{rev_r}-'P&L'!{m_c}{cost_r}")
    apply_style(ws_saas.cell(row=r, column=3 + i), font=font_total, number_format=NUM_DOLLAR, border=border_top_double)
r += 2

# Comparable Companies
style_section_header(ws_saas, r, "COMPARABLE COMPANIES", 7)
r += 1

for i, h in enumerate(["Company", "Category", "Est. ARR", "Notes"], 2):
    c = ws_saas.cell(row=r, column=i, value=h)
    apply_style(c, font=font_white_bold, fill=fill_light_navy, alignment=align_center)
r += 1

comps = [
    ("Chainalysis", "Blockchain Analytics", "$200M+", "Government & enterprise compliance data"),
    ("Nansen", "On-Chain Analytics", "$30M+", "Wallet labeling, smart money tracking"),
    ("Dune Analytics", "Data Dashboards", "$20M+", "Community-driven blockchain queries"),
    ("Gauntlet", "DeFi Risk Modeling", "$15M+", "Protocol risk simulations"),
    ("The Arena (Y2)", "AI Agent Intelligence", f"='P&L'!{mcol(24)}{pl_rows['SaaS ARR']}", "Agent Reliability Index — on-chain AI performance data"),
]

for name, cat, arr, notes in comps:
    ws_saas.cell(row=r, column=2, value=name).font = font_formula_bold
    ws_saas.cell(row=r, column=3, value=cat).font = font_label
    c_arr = ws_saas.cell(row=r, column=4, value=arr)
    if arr.startswith("="):
        apply_style(c_arr, font=font_formula_bold, number_format=NUM_DOLLAR)
    else:
        c_arr.font = font_formula_bold
    ws_saas.cell(row=r, column=5, value=notes).font = font_label
    r += 1


# ═══════════════════════════════════════════════════
# TAB 5: SCENARIOS
# ═══════════════════════════════════════════════════

ws_sc = wb.create_sheet("Scenarios")
ws_sc.sheet_properties.tabColor = "2C3E6B"
ws_sc.sheet_view.showGridLines = False

ws_sc.column_dimensions['A'].width = 3
ws_sc.column_dimensions['B'].width = 32
ws_sc.column_dimensions['C'].width = 18
ws_sc.column_dimensions['D'].width = 18
ws_sc.column_dimensions['E'].width = 18

# Title
ws_sc.merge_cells('A1:E1')
c = ws_sc.cell(row=1, column=1, value="SCENARIO ANALYSIS — BEAR / BASE / BULL")
apply_style(c, font=font_title, fill=fill_navy, alignment=align_center)
for cc in range(2, 6):
    apply_style(ws_sc.cell(row=1, column=cc), fill=fill_navy)

r = 3
style_section_header(ws_sc, r, "SCENARIO ASSUMPTIONS", 5)
r += 1

# Headers
for i, h in enumerate(["Assumption", "Bear", "Base", "Bull"], 2):
    c = ws_sc.cell(row=r, column=i, value=h)
    apply_style(c, font=font_white_bold, fill=fill_light_navy, alignment=align_center)
r += 1

scenario_assumptions = [
    ("Month 1 Tasks", 10, f"={a('Month 1 Spot Tasks')}", 50, NUM_INT),
    ("MoM Growth Rate", 0.15, f"={a('Spot Task MoM Growth Rate')}", 0.50, NUM_PCT),
    ("Avg Bounty (USD)", 800, f"={a('Average Bounty (USD)')}", 3000, NUM_DOLLAR),
    ("Bounty Growth", 0.02, f"={a('Bounty MoM Growth Rate')}", 0.08, NUM_PCT),
    ("Failure Rate", 0.30, f"={a('Failure Rate')}", 0.10, NUM_PCT),
    ("CC Retention", 0.80, f"={a('CC Retention Rate')}", 0.95, NUM_PCT),
    ("SaaS Growth", 0.10, f"={a('SaaS MoM Growth Rate')}", 0.35, NUM_PCT),
    ("Insurance Adoption", 0.08, f"={a('Insurance Adoption Rate')}", 0.25, NUM_PCT),
]

for label, bear, base, bull, fmt in scenario_assumptions:
    ws_sc.cell(row=r, column=2, value=label).font = font_label
    ws_sc.cell(row=r, column=3, value=bear)
    apply_style(ws_sc.cell(row=r, column=3), font=font_input, fill=fill_input, number_format=fmt, alignment=align_right)
    ws_sc.cell(row=r, column=4, value=base)
    apply_style(ws_sc.cell(row=r, column=4), font=font_formula, number_format=fmt, alignment=align_right)
    ws_sc.cell(row=r, column=5, value=bull)
    apply_style(ws_sc.cell(row=r, column=5), font=font_input, fill=fill_input, number_format=fmt, alignment=align_right)
    r += 1

r += 1

# Base Case Outputs (live from P&L)
style_section_header(ws_sc, r, "BASE CASE OUTPUTS (Live from P&L)", 5)
r += 1

for i, h in enumerate(["Metric", "M12", "M24"], 2):
    c = ws_sc.cell(row=r, column=i, value=h)
    apply_style(c, font=font_white_bold, fill=fill_light_navy, alignment=align_center)
r += 1

base_outputs = [
    ("Monthly Tasks", "Total Tasks", NUM_INT),
    ("Monthly GMV", "TOTAL GMV", NUM_DOLLAR),
    ("Monthly Revenue", "TOTAL REVENUE", NUM_DOLLAR),
    ("Net Income", "NET INCOME", NUM_DOLLAR),
    ("Cumulative P&L", "Cumulative P&L", NUM_DOLLAR),
    ("SaaS ARR", "SaaS ARR", NUM_DOLLAR),
    ("Total ARR", "Total ARR", NUM_DOLLAR),
    ("Take Rate", "Take Rate", NUM_PCT),
    ("Net Margin", "Net Margin", NUM_PCT),
    ("Cumulative Tasks", "Cumulative Tasks", NUM_INT),
    ("Cumulative GMV", "Cumulative GMV", NUM_DOLLAR),
]

for label, pl_label, fmt in base_outputs:
    ws_sc.cell(row=r, column=2, value=label).font = font_label

    # M12
    m12_r = pl_rows[pl_label]
    m12_c = mcol(12)
    ws_sc.cell(row=r, column=3, value=f"='P&L'!{m12_c}{m12_r}")
    apply_style(ws_sc.cell(row=r, column=3), font=font_formula, number_format=fmt, alignment=align_right)

    # M24
    m24_r = pl_rows[pl_label]
    m24_c = mcol(24)
    ws_sc.cell(row=r, column=4, value=f"='P&L'!{m24_c}{m24_r}")
    apply_style(ws_sc.cell(row=r, column=4), font=font_formula, number_format=fmt, alignment=align_right)

    r += 1


# ═══════════════════════════════════════════════════
# TAB 6: USE OF FUNDS
# ═══════════════════════════════════════════════════

ws_uof = wb.create_sheet("Use of Funds")
ws_uof.sheet_properties.tabColor = "2C3E6B"
ws_uof.sheet_view.showGridLines = False

ws_uof.column_dimensions['A'].width = 3
ws_uof.column_dimensions['B'].width = 35
ws_uof.column_dimensions['C'].width = 18
ws_uof.column_dimensions['D'].width = 14
ws_uof.column_dimensions['E'].width = 40

# Title
ws_uof.merge_cells('A1:E1')
c = ws_uof.cell(row=1, column=1, value="USE OF FUNDS — $250K PRE-SEED RAISE")
apply_style(c, font=font_title, fill=fill_navy, alignment=align_center)
for cc in range(2, 6):
    apply_style(ws_uof.cell(row=1, column=cc), fill=fill_navy)

r = 3
style_section_header(ws_uof, r, "CAPITAL ALLOCATION", 5)
r += 1

# Headers
for i, h in enumerate(["Category", "Amount", "% of Raise", "Purpose"], 2):
    c = ws_uof.cell(row=r, column=i, value=h)
    apply_style(c, font=font_white_bold, fill=fill_light_navy, alignment=align_center)
r += 1

total_raise = 250000
allocations = [
    ("Smart Contract Audit", 15000, "Formal verification + third-party audit"),
    ("Legal & Compliance", 10000, "Entity formation, regulatory review, ToS"),
    ("Seed Task Bounties", 20000, "Bootstrap initial protocol activity"),
    ("Agent Incentives", 15000, "Early agent onboarding rewards"),
    ("Infrastructure", 10000, "RPC nodes, IPFS, monitoring, hosting"),
    ("SaaS Development", 30000, "Agent Reliability Index data product build"),
    ("Marketing & Growth", 40000, "Community building, content, partnerships"),
    ("Bug Bounty Program", 20000, "Security researcher rewards"),
    ("Team (6 months)", 60000, "Core team compensation runway"),
    ("Reserve", 30000, "Operating buffer and contingency"),
]

alloc_start_row = r
for cat, amt, purpose in allocations:
    ws_uof.cell(row=r, column=2, value=cat).font = font_label
    ws_uof.cell(row=r, column=3, value=amt)
    apply_style(ws_uof.cell(row=r, column=3), font=font_input, fill=fill_input, number_format=NUM_DOLLAR)
    ws_uof.cell(row=r, column=4, value=f"=C{r}/{total_raise}")
    apply_style(ws_uof.cell(row=r, column=4), font=font_formula, number_format=NUM_PCT)
    ws_uof.cell(row=r, column=5, value=purpose).font = font_label
    r += 1

# Total
ws_uof.cell(row=r, column=2, value="TOTAL").font = font_total
ws_uof.cell(row=r, column=3, value=f"=SUM(C{alloc_start_row}:C{r-1})")
apply_style(ws_uof.cell(row=r, column=3), font=font_total, number_format=NUM_DOLLAR, border=border_top_double)
ws_uof.cell(row=r, column=4, value=f"=SUM(D{alloc_start_row}:D{r-1})")
apply_style(ws_uof.cell(row=r, column=4), font=font_total, number_format=NUM_PCT, border=border_top_double)

r += 2

# Milestones
style_section_header(ws_uof, r, "MILESTONES", 5)
r += 1

for i, h in enumerate(["Timeline", "Milestone", "Key Deliverables"], 2):
    c = ws_uof.cell(row=r, column=i, value=h)
    apply_style(c, font=font_white_bold, fill=fill_light_navy, alignment=align_center)
r += 1

milestones = [
    ("M1-2", "Testnet + Audit", "Smart contract audit complete, testnet live with seed tasks"),
    ("M3-4", "Mainnet + 50 Tasks", "Mainnet launch, 50+ tasks posted, first agent earnings"),
    ("M6", "SaaS Revenue", "Agent Reliability Index MVP launched, first paying subscribers"),
    ("M9", "CC + Insurance Live", "Continuous contracts active, insurance pool operational"),
    ("M12", "1,000 Tasks + Series A Ready", "1,000+ cumulative tasks, proven unit economics, fundraise deck"),
    ("M18", "Multi-Chain + SaaS ARR > Protocol Rev", "Expanded to 2+ chains, SaaS ARR exceeds protocol fee revenue"),
]

for timeline, milestone, deliverables in milestones:
    ws_uof.cell(row=r, column=2, value=timeline).font = font_formula_bold
    ws_uof.cell(row=r, column=3, value=milestone).font = font_label
    ws_uof.cell(row=r, column=4, value=deliverables).font = font_label
    ws_uof.merge_cells(f'D{r}:E{r}')
    ws_uof.cell(row=r, column=4).alignment = align_wrap
    ws_uof.row_dimensions[r].height = 30
    r += 1


# ═══════════════════════════════════════════════════
# FINAL: SAVE
# ═══════════════════════════════════════════════════

OUTPUT_PATH = "/Users/JackArnot/Desktop/The Arena Working /arena-codex/arena-financial-model.xlsx"
wb.save(OUTPUT_PATH)
print(f"\nSaved to: {OUTPUT_PATH}")
print(f"Sheets: {wb.sheetnames}")
print(f"P&L rows used: {current_row}")
print(f"Assumptions cells mapped: {len(assumptions)}")
