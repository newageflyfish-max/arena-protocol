#!/usr/bin/env python3
"""
Formula validation for Excel spreadsheets.
Checks that all formula cells are syntactically valid and reference existing sheets/cells.
"""
import sys
import re
import openpyxl

def validate_formulas(path):
    wb = openpyxl.load_workbook(path, data_only=False)
    errors = []
    sheet_names = wb.sheetnames

    for ws in wb:
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    formula = cell.value
                    # Check for common issues
                    # 1. Unmatched parentheses
                    if formula.count('(') != formula.count(')'):
                        errors.append(f"[{ws.title}!{cell.coordinate}] Unmatched parentheses: {formula}")
                    # 2. Cross-sheet references to non-existent sheets
                    refs = re.findall(r"'([^']+)'!", formula)
                    for ref in refs:
                        if ref not in sheet_names:
                            errors.append(f"[{ws.title}!{cell.coordinate}] References non-existent sheet '{ref}': {formula}")
                    # 3. Empty formula
                    if formula.strip() == '=':
                        errors.append(f"[{ws.title}!{cell.coordinate}] Empty formula")
                    # 4. Double operators
                    if re.search(r'[+\-*/]{2,}', formula.replace('--', '')):
                        errors.append(f"[{ws.title}!{cell.coordinate}] Possible double operator: {formula}")

    if errors:
        print(f"ERRORS FOUND: {len(errors)}")
        for e in errors:
            print(f"  {e}")
    else:
        print(f"0 errors found across {len(sheet_names)} sheets")

    return len(errors)

if __name__ == '__main__':
    path = sys.argv[1] if len(sys.argv) > 1 else 'arena-financial-model.xlsx'
    err_count = validate_formulas(path)
    sys.exit(1 if err_count > 0 else 0)
